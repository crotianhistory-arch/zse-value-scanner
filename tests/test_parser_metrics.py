import math
from zse_tool.parsers import parse_xlsx_report
from zse_tool.metrics import calculate_metrics
from zse_tool.validation import validate_report


def test_parser_and_metrics(sample_xlsx):
    r=parse_xlsx_report(sample_xlsx)
    assert r.issuer_name=="TESTCO d.d."
    assert r.year==2025 and r.quarter==1 and r.consolidated is True and r.audited is False
    assert r.period_end.isoformat()=="2025-03-31"
    assert validate_report(r)==[]
    m={x.name:x.value for x in calculate_metrics(r)}
    assert m["cash"]==20
    assert m["gross_financial_debt_ex_other"]==40
    assert m["net_debt_ex_other"]==20
    assert m["sales_revenue_ytd"]==50
    assert m["ebit_ytd"]==15
    assert m["ebitda_simple_ytd"]==20
    assert m["free_cash_flow_ytd"]==8
    assert math.isclose(m["net_debt_to_ebitda_run_rate"],0.25)
    assert math.isclose(m["interest_coverage_ebit"],7.5)
    assert math.isclose(m["roe_run_rate"],32/47.5)


def test_parser_accepts_croatian_eho_sheet_names(sample_xlsx, tmp_path):
    from openpyxl import load_workbook
    wb = load_workbook(sample_xlsx)
    wb["General data"].title = "Opći podaci"
    wb["Balance sheet"].title = "Bilanca"
    wb["P&L"].title = "RDG"
    wb["CF_D"].title = "NT_D"
    wb["CF_I"].title = "NT_I"
    g = wb["Opći podaci"]
    g["B6"] = "Godina:"
    g["B8"] = "Kvartal:"
    g["A19"] = "Tvrtka izdavatelja:"
    g["A31"] = "Konsolidirani izvještaj:"
    g["A33"] = "Revidirano:"
    p = tmp_path / "testco_hr.xlsx"
    wb.save(p)

    r = parse_xlsx_report(p)
    assert r.issuer_name == "TESTCO d.d."
    assert r.year == 2025 and r.quarter == 1
    assert r.consolidated is True and r.audited is False
    assert not any("does not balance" in w or "Cash mismatch" in w for w in validate_report(r))


def test_parser_ignores_eho_column_numbering_rows(sample_xlsx, tmp_path):
    from collections import Counter
    from openpyxl import load_workbook

    wb = load_workbook(sample_xlsx)

    # Real EHO templates include structural rows like:
    #   1 | ... | 2 | 3 | 4
    # These are column numbers, not ADP facts.
    bs = wb["Balance sheet"]
    bs.cell(6, 1, 1); bs.cell(6, 7, 2); bs.cell(6, 8, 3); bs.cell(6, 9, 4)

    pl = wb["P&L"]
    pl.cell(6, 1, 1); pl.cell(6, 7, 2); pl.cell(6, 8, 3); pl.cell(6, 9, 4); pl.cell(6, 10, 5); pl.cell(6, 11, 6)

    cfd = wb["CF_D"]
    cfd.cell(6, 1, 1); cfd.cell(6, 7, 2); cfd.cell(6, 8, 3); cfd.cell(6, 9, 4)

    cfi = wb["CF_I"]
    cfi.cell(6, 1, 1); cfi.cell(6, 7, 2); cfi.cell(6, 8, 3); cfi.cell(6, 9, 4)

    p = tmp_path / "testco_with_eho_headers.xlsx"
    wb.save(p)

    r = parse_xlsx_report(p)
    keys = [(f.statement, f.adp_code, f.column_name, f.source_sheet) for f in r.facts]
    duplicates = [key for key, count in Counter(keys).items() if count > 1]
    assert duplicates == []

    # The fake column-numbering row must not create ADP 2 facts in the balance sheet.
    assert not any(
        f.statement == "balance_sheet" and f.adp_code == 2 and f.source_row == 6
        for f in r.facts
    )


def test_parser_can_use_optional_schema_mapper_for_unknown_sheet_names(sample_xlsx, tmp_path):
    from openpyxl import load_workbook

    wb = load_workbook(sample_xlsx)
    rename = {
        "General data": "Podaci o izvjestaju Grupe",
        "Balance sheet": "Izvještaj o financijskom položaju Grupe",
        "P&L": "Izvještaj o rezultatu poslovanja Grupe",
        "CF_D": "Novčani tok - direktna metoda Grupe",
        "CF_I": "Novčani tok - indirektna metoda Grupe",
    }
    for old, new in rename.items():
        wb[old].title = new
    p = tmp_path / "testco_unknown_sheet_names.xlsx"
    wb.save(p)

    mapping = {
        "Podaci o izvjestaju Grupe": "general_data",
        "Izvještaj o financijskom položaju Grupe": "balance_sheet",
        "Izvještaj o rezultatu poslovanja Grupe": "income_statement",
        "Novčani tok - direktna metoda Grupe": "cash_flow_direct",
        "Novčani tok - indirektna metoda Grupe": "cash_flow_indirect",
    }

    class DummyMapper:
        def map_sheet_name(self, name):
            return mapping.get(name)

    r = parse_xlsx_report(p, schema_mapper=DummyMapper())
    assert r.issuer_name == "TESTCO d.d."
    assert len(r.facts) > 0
    assert any("LLM mapped sheet" in w for w in r.warnings)
    assert not any("does not balance" in w or "Cash mismatch" in w for w in validate_report(r))


def test_parser_infers_hrk_for_pre_euro_period_when_currency_missing(sample_xlsx, tmp_path):
    from datetime import datetime
    from openpyxl import load_workbook

    wb = load_workbook(sample_xlsx)
    g = wb["General data"]
    g["E4"] = datetime(2022, 1, 1)
    g["H4"] = datetime(2022, 12, 31)
    g["E6"] = 2022
    g["E8"] = "4."
    for name in ("Balance sheet", "P&L", "CF_D", "CF_I"):
        wb[name]["A3"] = None
    p = tmp_path / "testco_2022_no_currency.xlsx"
    wb.save(p)

    r = parse_xlsx_report(p)
    assert r.currency == "HRK"
    assert all(f.unit == "HRK" for f in r.facts)
    assert any("inferred HRK" in w for w in r.warnings)


def test_explicit_eur_overrides_pre_2023_currency_fallback(sample_xlsx, tmp_path):
    from datetime import datetime
    from openpyxl import load_workbook

    wb = load_workbook(sample_xlsx)
    g = wb["General data"]
    g["E4"] = datetime(2022, 1, 1)
    g["H4"] = datetime(2022, 12, 31)
    g["E6"] = 2022
    g["E8"] = "4."
    # sample_xlsx statement headers explicitly contain "in EUR"
    p = tmp_path / "testco_2022_explicit_eur.xlsx"
    wb.save(p)

    r = parse_xlsx_report(p)
    assert r.currency == "EUR"
    assert all(f.unit == "EUR" for f in r.facts)
    assert not any("inferred HRK" in w for w in r.warnings)


def test_parser_recognizes_croatian_kuna_header(sample_xlsx, tmp_path):
    from datetime import datetime
    from openpyxl import load_workbook

    wb = load_workbook(sample_xlsx)
    g = wb["General data"]
    g["E4"] = datetime(2022, 1, 1)
    g["H4"] = datetime(2022, 12, 31)
    g["E6"] = 2022
    g["E8"] = "4."
    for name in ("Balance sheet", "P&L", "CF_D", "CF_I"):
        wb[name]["A3"] = "u kunama"
    p = tmp_path / "testco_2022_kunama.xlsx"
    wb.save(p)

    r = parse_xlsx_report(p)
    assert r.currency == "HRK"
    assert all(f.unit == "HRK" for f in r.facts)
