from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from zse_tool.metrics import calculate_metrics
from zse_tool.models import EhoItem, Metric, ParsedReport
from zse_tool.parsers import parse_xlsx_report
from zse_tool.reporting import audit_label, period_label, scope_label
from zse_tool.storage import Database


def _save_report(
    db: Database,
    path: Path,
    *,
    source_id: str,
    publish_date: str,
    consolidated: bool,
    audited: bool,
    metric_value: float,
    year: int = 2025,
    quarter: int | None = 1,
    period_end: date = date(2025, 3, 31),
):
    path.write_bytes(b"placeholder")
    url = f"https://eho.zse.hr/{source_id}.xlsx"
    item = EhoItem(
        source_id=source_id,
        variant="financialReports",
        issuer_code="TEST",
        issuer_name="TEST d.d.",
        title=None,
        publish_date=publish_date,
        item_link=None,
        raw={},
        document_urls=[url],
    )
    db.upsert_items([item])
    db.mark_downloaded(
        source_id=source_id,
        variant="financialReports",
        url=url,
        local_path=path,
        sha256="x",
        byte_size=1,
    )
    report = ParsedReport(
        source_path=path,
        issuer_name="TEST d.d.",
        period_start=date(year, 1, 1),
        period_end=period_end,
        year=year,
        quarter=quarter,
        consolidated=consolidated,
        audited=audited,
        currency="EUR",
        scale=1.0,
        facts=[],
        warnings=[],
    )
    db.save_parsed_report(
        report,
        source_id=source_id,
        issuer_code="TEST",
        metrics=[Metric("cash", metric_value, "EUR")],
    )


def test_parser_accepts_croatian_text_reporting_dates(sample_xlsx: Path, tmp_path: Path):
    p = tmp_path / "croatian_text_dates.xlsx"
    wb = load_workbook(sample_xlsx)
    wb["General data"].title = "Opći podaci"
    ws = wb["Opći podaci"]
    ws["E4"] = "1.1.2025."
    ws["H4"] = "31.03.2025."
    wb.save(p)

    parsed = parse_xlsx_report(p)
    assert parsed.period_start == date(2025, 1, 1)
    assert parsed.period_end == date(2025, 3, 31)


def test_period_scope_and_audit_labels():
    assert period_label(year=2025, quarter=1) == "2025-Q1"
    assert period_label(year=2024, quarter=None, period_end=date(2024, 12, 31)) == "2024-FY"
    assert scope_label(True) == "CONSOLIDATED"
    assert scope_label(False) == "UNCONSOLIDATED"
    assert audit_label(True) == "AUDITED"


def test_report_inventory_prefers_consolidated(tmp_path: Path):
    db = Database(tmp_path / "x.sqlite")
    _save_report(
        db,
        tmp_path / "unconsolidated.xlsx",
        source_id="u",
        publish_date="2025-04-30 12:00",
        consolidated=False,
        audited=False,
        metric_value=10.0,
    )
    _save_report(
        db,
        tmp_path / "consolidated.xlsx",
        source_id="c",
        publish_date="2025-04-29 12:00",
        consolidated=True,
        audited=False,
        metric_value=20.0,
    )

    rows = db.report_inventory("TEST")
    assert rows[0]["source_id"] == "c"
    assert rows[0]["preference_rank"] == 1
    assert rows[0]["period_candidates"] == 2


def test_report_inventory_prefers_newest_same_scope_as_correction(tmp_path: Path):
    db = Database(tmp_path / "x.sqlite")
    _save_report(
        db,
        tmp_path / "old.xlsx",
        source_id="old",
        publish_date="2025-04-29 12:00",
        consolidated=True,
        audited=False,
        metric_value=20.0,
    )
    _save_report(
        db,
        tmp_path / "new.xlsx",
        source_id="new",
        publish_date="2025-05-02 12:00",
        consolidated=True,
        audited=False,
        metric_value=21.0,
    )

    rows = db.report_inventory("TEST", preferred_only=True)
    assert len(rows) == 1
    assert rows[0]["source_id"] == "new"


def test_latest_metrics_uses_preferred_report(tmp_path: Path):
    db = Database(tmp_path / "x.sqlite")
    _save_report(
        db,
        tmp_path / "unconsolidated.xlsx",
        source_id="u",
        publish_date="2025-04-30 12:00",
        consolidated=False,
        audited=False,
        metric_value=10.0,
    )
    _save_report(
        db,
        tmp_path / "consolidated.xlsx",
        source_id="c",
        publish_date="2025-04-29 12:00",
        consolidated=True,
        audited=False,
        metric_value=20.0,
    )

    rows = db.latest_metrics("TEST")
    cash = next(r for r in rows if r["metric_name"] == "cash")
    assert cash["value"] == 20.0
    assert cash["consolidated"] == 1
