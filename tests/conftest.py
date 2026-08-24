from pathlib import Path
import pytest
from openpyxl import Workbook


@pytest.fixture
def sample_xlsx(tmp_path: Path) -> Path:
    p=tmp_path/"testco_q1.xlsx"
    wb=Workbook(); wb.remove(wb.active)
    g=wb.create_sheet("General data")
    g["A4"]="Reporting period:"; g["E4"]="2025-01-01"; g["H4"]="2025-03-31"
    # dates must be native dates for the parser
    from datetime import datetime
    g["E4"]=datetime(2025,1,1); g["H4"]=datetime(2025,3,31)
    g["B6"]="Year:"; g["E6"]=2025
    g["B8"]="Quarter:"; g["E8"]="1."
    g["A19"]="Name of the issuer:"; g["C19"]="TESTCO d.d."
    g["A31"]="Consolidated report:"; g["C31"]="KD"
    g["A33"]="Audited:"; g["C33"]="RN"

    def base_sheet(name, title):
        ws=wb.create_sheet(name); ws["A1"]=title; ws["A3"]="in EUR"; ws["G5"]="ADP code"; return ws
    bs=base_sheet("Balance sheet","BALANCE SHEET")
    rows=[
      (63,"CASH",15,20),(65,"TOTAL ASSETS",95,100),(67,"CAPITAL AND RESERVES",50,60),(89,"NCI",5,10),
      (99,"LT group loans",0,0),(101,"LT related loans",0,0),(102,"LT loans",0,0),(103,"LT banks",20,30),(106,"LT securities",0,0),
      (111,"ST group loans",0,0),(113,"ST related loans",0,0),(114,"ST loans",0,0),(115,"ST banks",5,10),(118,"ST securities",0,0),(125,"TOTAL",95,100)]
    for i,(code,label,prev,cur) in enumerate(rows,8): bs.cell(i,1,label); bs.cell(i,7,code); bs.cell(i,8,prev); bs.cell(i,9,cur)

    pl=base_sheet("P&L","P&L")
    rows=[(1,"OPERATING INCOME",40,40,55,55),(2,"group sales",0,0,0,0),(3,"sales",45,45,50,50),(7,"OPERATING EXPENSES",30,30,40,40),(17,"Depreciation",4,4,5,5),(42,"group interest",0,0,0,0),(44,"interest",1,1,2,2),(75,"NET INCOME",7,7,10,10),(76,"parent NI",6,6,8,8)]
    for i,(code,label,a,b,c,d) in enumerate(rows,8):
        pl.cell(i,1,label); pl.cell(i,7,code); pl.cell(i,8,a); pl.cell(i,9,b); pl.cell(i,10,c); pl.cell(i,11,d)

    cfd=base_sheet("CF_D","CASH FLOW DIRECT")
    for i,(code,label,prev,cur) in enumerate([(14,"CFO",10,12),(22,"CAPEX",-3,-4),(44,"CASH END",15,20)],8):
        cfd.cell(i,1,label); cfd.cell(i,7,code); cfd.cell(i,8,prev); cfd.cell(i,9,cur)
    cfi=base_sheet("CF_I","CASH FLOW INDIRECT")
    for i,(code,label) in enumerate([(20,"CFO"),(28,"CAPEX"),(50,"CASH END")],8):
        cfi.cell(i,1,label); cfi.cell(i,7,code); cfi.cell(i,8,0); cfi.cell(i,9,0)
    wb.create_sheet("SOCE"); wb.create_sheet("Notes")
    wb.save(p); return p
