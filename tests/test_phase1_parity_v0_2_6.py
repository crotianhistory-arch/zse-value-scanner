
from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook

from zse_tool.analytics import build_ttm_snapshot
from zse_tool.metrics import calculate_metrics
from zse_tool.models import EhoItem, Fact, Metric, ParsedReport
from zse_tool.parsers.xlsx_financial import _extract_employee_facts
from zse_tool.storage import Database


def _report(facts: list[Fact], *, consolidated: bool = True) -> ParsedReport:
    return ParsedReport(
        source_path=Path("test.xlsx"),
        issuer_name="TEST",
        period_start=date(2026, 1, 1),
        period_end=date(2026, 3, 31),
        year=2026,
        quarter=1,
        consolidated=consolidated,
        audited=False,
        currency="EUR",
        scale=1.0,
        facts=facts,
        warnings=[],
    )


def test_employee_notes_extract_average_group_count():
    wb = Workbook()
    ws = wb.active
    ws.title = "Bilješke"
    ws["A1"] = "6. Prosječan broj zaposlenih tijekom tekućeg razdoblja - Grupa"
    ws["B1"] = 5610

    facts = _extract_employee_facts(wb, consolidated=True)
    got = {(f.column_name, f.value, f.unit) for f in facts}
    assert ("employees_average_group", 5610.0, "count") in got


def test_employee_notes_keep_average_and_period_end_separate():
    wb = Workbook()
    ws = wb.active
    ws.title = "Notes"
    ws["A1"] = "Average number of employees during the current period - Group"
    ws["B1"] = "5,610"
    ws["A2"] = "Number of employees at period end - Group"
    ws["B2"] = "5,640"

    facts = _extract_employee_facts(wb, consolidated=True)
    by_col = {f.column_name: f.value for f in facts}
    assert by_col["employees_average_group"] == 5610.0
    assert by_col["employees_period_end_group"] == 5640.0


def test_metrics_preserve_both_employee_measures_and_summary_prefers_period_end():
    facts = [
        Fact("supplemental", 1, "avg", "employees_average_group", 5610.0, "count", "Bilješke", 10),
        Fact("supplemental", 1, "end", "employees_period_end_group", 5640.0, "count", "Bilješke", 11),
    ]
    metrics = {m.name: m for m in calculate_metrics(_report(facts))}
    assert metrics["employees_average_current_period"].value == 5610.0
    assert metrics["employees_period_end"].value == 5640.0
    assert metrics["employees_reported"].value == 5640.0
    assert metrics["employees_reported"].unit == "count"
    assert "period-end" in (metrics["employees_reported"].note or "")


def _save_metrics(
    db: Database,
    root: Path,
    sid: str,
    year: int,
    quarter: int | None,
    end: date,
    values: dict[str, tuple[float, str] | float],
    *,
    audited: bool = False,
):
    path = root / f"{sid}.xlsx"
    path.write_bytes(b"x")
    url = f"https://eho.zse.hr/{sid}.xlsx"
    db.upsert_items([
        EhoItem(
            source_id=sid,
            variant="financialReports",
            issuer_code="TEST",
            issuer_name="TEST",
            title=None,
            publish_date=end.isoformat(),
            item_link=None,
            raw={},
            document_urls=[url],
        )
    ])
    db.mark_downloaded(
        source_id=sid,
        variant="financialReports",
        url=url,
        local_path=path,
        sha256=sid,
        byte_size=1,
    )
    report = ParsedReport(
        source_path=path,
        issuer_name="TEST",
        period_start=date(year, 1, 1),
        period_end=end,
        year=year,
        quarter=quarter,
        consolidated=True,
        audited=audited,
        currency="EUR",
        scale=1.0,
        facts=[],
        warnings=[],
    )
    metrics = []
    for name, raw in values.items():
        if isinstance(raw, tuple):
            value, unit = raw
        else:
            value, unit = raw, "EUR"
        metrics.append(Metric(name, value, unit))
    db.save_parsed_report(
        report,
        source_id=sid,
        source_variant="financialReports",
        source_url=url,
        source_publish_date=end.isoformat(),
        source_sha256=sid,
        issuer_code="TEST",
        metrics=metrics,
    )


def _base_money():
    # Minimum set used by snapshot ratios. Missing unrelated values are fine.
    return {
        "sales_revenue_ytd": 100.0,
        "operating_revenue_ytd": 100.0,
        "total_revenue_ytd": 100.0,
        "ebit_ytd": 20.0,
        "depreciation_ytd": 10.0,
        "value_adjustments_ytd": 0.0,
        "provisions_ytd": 0.0,
        "md_value_adjustments_provisions_ytd": 0.0,
        "ebitda_simple_ytd": 30.0,
        "financial_income_ytd": 1.0,
        "financial_expense_ytd": 2.0,
        "ebt_ytd": 19.0,
        "md_ebit_ytd": 20.0,
        "md_ebitda_ytd": 30.0,
        "net_income_parent_ytd": 15.0,
        "net_income_total_ytd": 15.0,
        "interest_expense_ytd": 1.0,
        "operating_cash_flow_ytd": 20.0,
        "asset_sale_proceeds_ytd": 0.0,
        "capex_ytd": 5.0,
        "net_capex_ytd": 5.0,
        "free_cash_flow_ytd": 15.0,
        "free_cash_flow_net_capex_ytd": 15.0,
        "cash": 10.0,
        "short_term_financial_assets": 0.0,
        "liquid_financial_assets": 10.0,
        "current_assets": 50.0,
        "current_liabilities": 20.0,
        "long_term_liabilities": 10.0,
        "total_debt_md": 30.0,
        "total_assets": 100.0,
        "equity_total": 60.0,
        "equity_nci": 0.0,
        "equity_parent": 60.0,
        "share_capital": 20.0,
        "retained_earnings_reserves_parent": 40.0,
        "gross_financial_debt_ex_other": 5.0,
        "gross_financial_debt_standardized": 5.0,
        "net_debt_ex_other": -5.0,
        "net_debt_liquid_assets": -5.0,
    }


def test_q1_comprehensive_ttm_uses_redundant_quarter_column_when_available(tmp_path: Path):
    db = Database(tmp_path / "db.sqlite")

    fy = _base_money()
    fy["comprehensive_income_parent_ytd"] = 144.20
    _save_metrics(db, tmp_path, "fy25", 2025, None, date(2025, 12, 31), fy, audited=True)

    q125 = _base_money()
    q125["comprehensive_income_parent_ytd"] = 144.20  # simulate bad cumulative cell
    q125["comprehensive_income_parent_quarter"] = 38.52
    _save_metrics(db, tmp_path, "q125", 2025, 1, date(2025, 3, 31), q125)

    q126 = _base_money()
    q126["comprehensive_income_parent_ytd"] = 40.73
    q126["comprehensive_income_parent_quarter"] = 40.73
    q126["employees_average_current_period"] = (5610.0, "count")
    q126["employees_reported"] = (5610.0, "count")
    _save_metrics(db, tmp_path, "q126", 2026, 1, date(2026, 3, 31), q126)

    snap = build_ttm_snapshot(db, "TEST", "2026-Q1")
    comp = snap["metrics"]["comprehensive_income_parent_ttm"]
    assert comp["value"] == pytest.approx(146.41)
    assert comp["sources"][0]["metric"] == "comprehensive_income_parent_quarter"
    assert snap["metrics"]["employees_reported"]["value"] == pytest.approx(5610.0)
    assert snap["metrics"]["employees_reported"]["unit"] == "count"
