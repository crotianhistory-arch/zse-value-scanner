from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
import re
import unicodedata
from typing import Iterable

from openpyxl import load_workbook

from ..errors import UnsupportedReport
from ..models import Fact, ParsedReport

GENERAL_SHEET_ALIASES = ("General data", "Opći podaci", "Opci podaci")
KNOWN_NON_CORE_SHEETS = ("SOCE", "PK", "Notes", "Bilješke", "Biljeske")

SHEET_SPECS = {
    "Balance sheet": {
        "aliases": ("Balance sheet", "Bilanca"),
        "statement": "balance_sheet",
        "columns": {8: "previous_period", 9: "current_period"},
    },
    "P&L": {
        "aliases": ("P&L", "RDG"),
        "statement": "income_statement",
        "columns": {
            8: "previous_cumulative",
            9: "previous_quarter",
            10: "current_cumulative",
            11: "current_quarter",
        },
    },
    "CF_D": {
        "aliases": ("CF_D", "NT_D"),
        "statement": "cash_flow_direct",
        "columns": {8: "previous_period", 9: "current_period"},
    },
    "CF_I": {
        "aliases": ("CF_I", "NT_I"),
        "statement": "cash_flow_indirect",
        "columns": {8: "previous_period", 9: "current_period"},
    },
}


def _to_float(value) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().replace(" ", "")
        if not cleaned:
            return None
        cleaned = cleaned.replace(".", "") if cleaned.count(",") == 1 and cleaned.count(".") > 0 else cleaned
        cleaned = cleaned.replace(",", ".")
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _parse_currency_and_scale(ws, *, fallback_currency: str = "EUR") -> tuple[str, float, bool]:
    """Return statement currency, scale and whether currency was explicit.

    Migrated EHO workbooks from the kuna era often omit a machine-readable
    currency marker in the statement header.  Never default those silently to
    EUR: callers provide a reporting-period fallback (HRK through 2022, EUR
    from 2023 onward).  An explicit workbook marker always wins.
    """
    currency: str | None = None
    scale = 1.0
    max_row = min(ws.max_row, 12)
    max_col = min(ws.max_column, 6)
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            value = ws.cell(row, col).value
            if not isinstance(value, str):
                continue
            lower = value.casefold()
            if currency is None:
                if re.search(r"\bhrk\b", lower) or "kuna" in lower or "kunama" in lower:
                    currency = "HRK"
                elif re.search(r"\beur\b", lower) or "euro" in lower or "eurima" in lower:
                    currency = "EUR"
            if "thousand" in lower or "000" in lower or "tisuć" in lower:
                scale = 1000.0
    explicit = currency is not None
    return currency or fallback_currency, scale, explicit


def _fallback_currency_from_metadata(metadata: dict) -> str:
    """Infer Croatian filing currency only when the workbook does not state it.

    The reporting period controls the fallback, not the publication date.
    Croatian periods ending in 2022 or earlier are treated as HRK; periods
    from 2023 onward as EUR.  Explicit workbook currency markers override this.
    """
    period_end = metadata.get("period_end")
    if period_end is not None:
        return "HRK" if period_end.year <= 2022 else "EUR"
    year = metadata.get("year")
    if isinstance(year, int):
        return "HRK" if year <= 2022 else "EUR"
    return "EUR"


def _find_general_value(ws, label_fragment: str):
    needle = label_fragment.lower()
    for row in ws.iter_rows():
        for idx, cell in enumerate(row):
            if isinstance(cell.value, str) and needle in cell.value.lower():
                # Scan to the right for the first nonempty value.
                for later in row[idx + 1 :]:
                    if later.value not in (None, ""):
                        return later.value
    return None


def _find_general_value_aliases(ws, *label_fragments: str):
    for fragment in label_fragments:
        value = _find_general_value(ws, fragment)
        if value not in (None, ""):
            return value
    return None



def _employee_count_from_value(value) -> float | None:
    """Parse an explicitly reported employee count without guessing from money/date fields."""
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        v = float(value)
        if 0 <= v <= 10_000_000 and abs(v - round(v)) < 1e-6:
            return float(round(v))
        return None
    if not isinstance(value, str):
        return None
    s = value.strip()
    if not s:
        return None
    # Allow Croatian/European thousands separators (5.610 / 5 610 / 5,610)
    # but reject dates and arbitrary prose numbers by requiring one integer token.
    m = re.fullmatch(r"\s*(\d{1,3}(?:[ .,\u00a0]\d{3})+|\d{1,7})\s*", s)
    if not m:
        return None
    digits = re.sub(r"\D", "", m.group(1))
    if not digits:
        return None
    v = int(digits)
    return float(v) if 0 <= v <= 10_000_000 else None


def _employee_label_kind(text: str) -> tuple[str | None, str]:
    """Classify explicit employee disclosures."""
    norm = unicodedata.normalize("NFKD", text or "")
    norm = "".join(ch for ch in norm if not unicodedata.combining(ch)).casefold()
    norm = re.sub(r"\s+", " ", norm)
    if not (("zaposlen" in norm) or ("employee" in norm)):
        return None, "unspecified"
    if any(x in norm for x in ("trosak zaposlen", "employee cost", "staff cost", "placa", "salary", "wage")):
        return None, "unspecified"

    measure = None
    if ("prosjec" in norm or "average" in norm) and ("zaposlen" in norm or "employee" in norm):
        measure = "average"
    elif any(x in norm for x in (
        "na dan", "na 31", "na kraju", "krajem", "at 31", "as at", "at period end", "end of period",
        "year end", "year-end", "broj zaposlenih", "broj zaposlenika", "number of employees", "headcount",
    )):
        measure = "period_end"

    scope = "unspecified"
    if any(x in norm for x in ("grupa", "group", "koncern")):
        scope = "group"
    elif any(x in norm for x in ("drustv", "company", "parent company")):
        scope = "company"
    return measure, scope


def _extract_employee_facts(wb, *, consolidated: bool | None) -> list[Fact]:
    """Extract explicitly reported employee statistics from notes/general sheets.

    Croatian interim-report notes are required to disclose the average number
    of employees during the current period. Some issuers additionally disclose
    period-end headcount. Preserve measure and scope separately and never infer
    one from the other.
    """
    candidates: dict[tuple[str, str], tuple[int, int, float, str, int]] = {}
    scan_order = 0

    sheet_names: list[str] = []
    for name in wb.sheetnames:
        folded = unicodedata.normalize("NFKD", name)
        folded = "".join(ch for ch in folded if not unicodedata.combining(ch)).casefold()
        if name in GENERAL_SHEET_ALIASES or any(k in folded for k in ("biljes", "notes", "note")):
            sheet_names.append(name)

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        max_row = min(ws.max_row, 500)
        max_col = min(ws.max_column, 30)
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                raw = ws.cell(r, c).value
                if not isinstance(raw, str):
                    continue
                measure, scope = _employee_label_kind(raw)
                if measure is None:
                    continue
                scan_order += 1

                values: list[tuple[int, float]] = []
                for dc in range(1, 7):
                    if c + dc <= max_col:
                        v = _employee_count_from_value(ws.cell(r, c + dc).value)
                        if v is not None:
                            values.append((100 - dc, v))
                for dr in range(1, 3):
                    for dc in range(0, 4):
                        if r + dr <= max_row and c + dc <= max_col:
                            v = _employee_count_from_value(ws.cell(r + dr, c + dc).value)
                            if v is not None:
                                values.append((70 - 5 * dr - dc, v))

                if not values:
                    tail_numbers = re.findall(r"(?<!\d)(\d{1,3}(?:[ .,\u00a0]\d{3})+|\d{1,7})(?!\d)", raw)
                    for token in tail_numbers:
                        v = _employee_count_from_value(token)
                        if v is not None and not (1900 <= v <= 2100):
                            values.append((40, v))
                            break

                if not values:
                    continue
                local_score, value = max(values, key=lambda x: x[0])

                scope_bonus = 0
                if consolidated is True and scope == "group":
                    scope_bonus = 20
                elif consolidated is False and scope == "company":
                    scope_bonus = 20
                elif scope == "unspecified":
                    scope_bonus = 5

                key = (measure, scope)
                score = local_score + scope_bonus
                prev = candidates.get(key)
                if prev is None or score > prev[0] or (score == prev[0] and scan_order < prev[1]):
                    candidates[key] = (score, scan_order, value, sheet_name, r)

    facts: list[Fact] = []
    col_map = {
        ("average", "group"): "employees_average_group",
        ("average", "company"): "employees_average_company",
        ("average", "unspecified"): "employees_average_unspecified",
        ("period_end", "group"): "employees_period_end_group",
        ("period_end", "company"): "employees_period_end_company",
        ("period_end", "unspecified"): "employees_period_end_unspecified",
    }
    for key, (_score, _order, value, sheet_name, source_row) in candidates.items():
        facts.append(Fact(
            statement="supplemental",
            adp_code=1,
            label=f"reported employee count ({key[0]}, {key[1]})",
            column_name=col_map[key],
            value=value,
            unit="count",
            source_sheet=sheet_name,
            source_row=source_row,
        ))
    return facts


def _parse_dateish(value) -> date | None:
    """Parse native Excel dates and common EHO text-date variants.

    Some issuer workbooks store the reporting period as real Excel dates while
    others store strings such as ``1.1.2025.`` / ``31.03.2025.``.  Treat both
    deterministically; do not infer a date merely from year/quarter.
    """
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not isinstance(value, str):
        return None
    text = value.strip()
    if not text:
        return None
    for fmt in (
        "%d.%m.%Y.",
        "%d.%m.%Y",
        "%d.%m.%y.",
        "%d.%m.%y",
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
    ):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def _resolve_sheet_name(wb, aliases: Iterable[str]) -> str | None:
    # Exact first, then normalized case/whitespace match.
    for alias in aliases:
        if alias in wb.sheetnames:
            return alias
    normalized = {re.sub(r"\s+", " ", name).strip().casefold(): name for name in wb.sheetnames}
    for alias in aliases:
        key = re.sub(r"\s+", " ", alias).strip().casefold()
        if key in normalized:
            return normalized[key]
    return None


def _extract_general_metadata(wb, general_sheet: str | None = None) -> dict:
    general_sheet = general_sheet or _resolve_sheet_name(wb, GENERAL_SHEET_ALIASES)
    if general_sheet is None:
        return {}
    ws = wb[general_sheet]
    year = _find_general_value_aliases(ws, "Year:", "Godina:")
    quarter = _find_general_value_aliases(ws, "Quarter:", "Kvartal:")
    issuer = _find_general_value_aliases(ws, "Name of the issuer", "Tvrtka izdavatelja", "Naziv izdavatelja")
    consolidated_code = _find_general_value_aliases(ws, "Consolidated report", "Konsolidirani izvještaj", "Konsolidirani izvjestaj")
    audited_code = _find_general_value_aliases(ws, "Audited", "Revidirano")

    period_start = period_end = None
    for row in range(1, min(ws.max_row, 15) + 1):
        values = [ws.cell(row, col).value for col in range(1, min(ws.max_column, 12) + 1)]
        dates = [d for d in (_parse_dateish(v) for v in values) if d is not None]
        if len(dates) >= 2:
            period_start, period_end = dates[0], dates[1]
            break

    def int_from(value):
        if isinstance(value, (int, float)):
            return int(value)
        if isinstance(value, str):
            m = re.search(r"\d+", value)
            return int(m.group()) if m else None
        return None

    consolidated = None
    if isinstance(consolidated_code, str):
        upper = consolidated_code.strip().upper()
        if upper == "KD":
            consolidated = True
        elif upper == "KN":
            consolidated = False

    audited = None
    if isinstance(audited_code, str):
        upper = audited_code.strip().upper()
        if upper == "RD":
            audited = True
        elif upper == "RN":
            audited = False

    return {
        "issuer_name": str(issuer).strip() if issuer not in (None, "") else None,
        "period_start": period_start,
        "period_end": period_end,
        "year": int_from(year),
        "quarter": int_from(quarter),
        "consolidated": consolidated,
        "audited": audited,
    }


def _parse_statement_sheet(ws, statement: str, column_map: dict[int, str], currency: str, scale: float) -> list[Fact]:
    facts: list[Fact] = []
    for row_idx in range(1, ws.max_row + 1):
        label = ws.cell(row_idx, 1).value
        adp = ws.cell(row_idx, 7).value
        adp_num = _to_float(adp)
        if adp_num is None or int(adp_num) < 0:
            continue
        adp_code = int(adp_num)
        if adp_code == 0 and not isinstance(label, str):
            continue
        label_text = str(label).strip() if label not in (None, "") else ""

        # EHO financial statement templates contain a column-numbering header row
        # such as 1 | ... | 2 | 3 | 4 (or 1..6 in the P&L).  The ADP column in
        # that row therefore looks numeric and used to be misread as a real fact,
        # producing duplicate ADP codes (notably ADP 2).  Real line-item labels
        # contain descriptive text (e.g. "1. Prihodi ..."), so a label consisting
        # only of an integer / optional trailing dot is a structural header, not data.
        if not label_text or re.fullmatch(r"\d+\.?", label_text):
            continue

        for col_idx, col_name in column_map.items():
            value = _to_float(ws.cell(row_idx, col_idx).value)
            if value is not None:
                value *= scale
            facts.append(
                Fact(
                    statement=statement,
                    adp_code=adp_code,
                    label=label_text,
                    column_name=col_name,
                    value=value,
                    unit=currency,
                    source_sheet=ws.title,
                    source_row=row_idx,
                )
            )
    return facts


def _assert_unique_fact_keys(facts: list[Fact], path: Path) -> None:
    """Fail early with a useful parser error instead of a SQLite UNIQUE error."""
    seen: dict[tuple[str, int, str, str], Fact] = {}
    for fact in facts:
        key = (fact.statement, fact.adp_code, fact.column_name, fact.source_sheet)
        previous = seen.get(key)
        if previous is not None:
            raise UnsupportedReport(
                "Duplicate financial fact after XLSX parsing in "
                f"{path.name}: statement={fact.statement}, ADP={fact.adp_code}, "
                f"column={fact.column_name}, sheet={fact.source_sheet}; "
                f"rows {previous.source_row} and {fact.source_row}. "
                "This usually means the EHO template structure changed and the parser needs an update."
            )
        seen[key] = fact


def _resolve_workbook_sheets(wb, schema_mapper=None) -> tuple[str | None, dict[str, str], list[str]]:
    general_sheet = _resolve_sheet_name(wb, GENERAL_SHEET_ALIASES)
    resolved: dict[str, str] = {}
    warnings: list[str] = []
    for canonical_name, spec in SHEET_SPECS.items():
        actual_name = _resolve_sheet_name(wb, spec.get("aliases", (canonical_name,)))
        if actual_name is not None:
            resolved[canonical_name] = actual_name

    if schema_mapper is None:
        return general_sheet, resolved, warnings

    semantic_to_canonical = {
        "general_data": "__general__",
        "balance_sheet": "Balance sheet",
        "income_statement": "P&L",
        "cash_flow_direct": "CF_D",
        "cash_flow_indirect": "CF_I",
    }
    already_used = set(resolved.values())
    if general_sheet:
        already_used.add(general_sheet)
    normalized_non_core = {re.sub(r"\s+", " ", x).strip().casefold() for x in KNOWN_NON_CORE_SHEETS}

    for actual_name in wb.sheetnames:
        if actual_name in already_used:
            continue
        if re.sub(r"\s+", " ", actual_name).strip().casefold() in normalized_non_core:
            continue
        mapped = schema_mapper.map_sheet_name(actual_name)
        target = semantic_to_canonical.get(mapped or "")
        if target == "__general__" and general_sheet is None:
            general_sheet = actual_name
            warnings.append(f"LLM mapped sheet {actual_name!r} -> general_data")
            already_used.add(actual_name)
        elif target in SHEET_SPECS and target not in resolved:
            resolved[target] = actual_name
            warnings.append(f"LLM mapped sheet {actual_name!r} -> {SHEET_SPECS[target]['statement']}")
            already_used.add(actual_name)

    return general_sheet, resolved, warnings


def parse_xlsx_report(path: str | Path, schema_mapper=None) -> ParsedReport:
    path = Path(path)
    if path.suffix.lower() != ".xlsx":
        raise UnsupportedReport("v0.1 parser supports .xlsx only; legacy .xls and ESEF come later")

    wb = load_workbook(path, data_only=True, read_only=True)
    general_sheet, resolved_sheets, mapping_warnings = _resolve_workbook_sheets(wb, schema_mapper)
    metadata = _extract_general_metadata(wb, general_sheet)
    facts: list[Fact] = []
    warnings: list[str] = list(mapping_warnings)

    currency = "EUR"
    scale = 1.0

    if not resolved_sheets:
        raise UnsupportedReport(
            f"No supported financial statement sheets found in {path.name}; "
            f"workbook sheets={wb.sheetnames}"
        )

    # Infer unit from the workbook when possible. Older migrated EHO files can
    # omit the currency marker entirely, so fall back from the reporting period
    # (HRK through 2022; EUR from 2023). Explicit sheet currency always wins.
    fallback_currency = _fallback_currency_from_metadata(metadata)
    first_actual = next(iter(resolved_sheets.values()))
    first_ws = wb[first_actual]
    currency, scale, currency_explicit = _parse_currency_and_scale(
        first_ws, fallback_currency=fallback_currency
    )
    if not currency_explicit:
        warnings.append(
            f"Currency not explicit in statement header; inferred {currency} "
            f"from reporting period."
        )

    for sheet_name, spec in SHEET_SPECS.items():
        actual_name = resolved_sheets.get(sheet_name)
        if actual_name is None:
            warnings.append(f"Missing sheet: {sheet_name} (aliases: {', '.join(spec.get('aliases', (sheet_name,)))})")
            continue
        ws = wb[actual_name]
        sheet_currency, sheet_scale, _ = _parse_currency_and_scale(
            ws, fallback_currency=fallback_currency
        )
        if sheet_currency != currency or sheet_scale != scale:
            warnings.append(
                f"Unit mismatch in {sheet_name}: {sheet_currency} x{sheet_scale}; "
                f"using sheet-specific unit for parsed facts"
            )
        facts.extend(
            _parse_statement_sheet(
                ws,
                spec["statement"],
                spec["columns"],
                sheet_currency,
                sheet_scale,
            )
        )

    facts.extend(_extract_employee_facts(wb, consolidated=metadata.get("consolidated")))

    _assert_unique_fact_keys(facts, path)

    return ParsedReport(
        source_path=path,
        issuer_name=metadata.get("issuer_name"),
        period_start=metadata.get("period_start"),
        period_end=metadata.get("period_end"),
        year=metadata.get("year"),
        quarter=metadata.get("quarter"),
        consolidated=metadata.get("consolidated"),
        audited=metadata.get("audited"),
        currency=currency,
        scale=scale,
        facts=facts,
        warnings=warnings,
    )
